from tkinter import *
from tkinter import ttk
import requests
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
from tkinter import filedialog


class TablePage:
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.selected_items = []

        self.create_widgets()
        self.load_test_results()

    def create_widgets(self):
        # Фрейм для верхних виджетов
        self.top_frame = Frame(self.parent, background='white')
        self.top_frame.pack(fill=X)

        # Надпись "Результаты ваших тестирований"
        self.label_YourResults = ttk.Label(self.top_frame,
                                           text='История результатов',
                                           font=('Arial', 14, 'bold'),
                                           style='Label.TLabel')
        self.label_YourResults.pack(side='left', padx=30, pady=10)

        # Надпись "Выберите строку для удаления"
        self.LableError = ttk.Label(self.top_frame,
                                    text="Выберите строку для удаления",
                                    font=('Calibri', 10, 'bold'),
                                    foreground='white',
                                    style='Label.TLabel')
        self.LableError.pack(side=RIGHT, padx=30, pady=10)

        # Фрейм для таблицы
        self.table_frame = Frame(self.parent,
                                 background='white',
                                 relief=SOLID,
                                 borderwidth=2)
        self.table_frame.pack(fill='both', expand=TRUE, padx=30, pady=20)

        self.create_table()

        # Кнопка "Назад", переход на главное меню
        self.BackBtn = ttk.Button(self.parent,
                                  text='Назад',
                                  command=self.back_button,
                                  style='StyleGray.TButton')
        self.BackBtn.pack(side='left', anchor='s', padx=30, pady=10, ipady=5)

        # Кнопка экспорта таблицы результатов в Excel
        self.ExportBtn = ttk.Button(self.parent,
                                    text='Экспорт результатов в Excel',
                                    style="StyleGreen.TButton",
                                    command=self.export_excel)
        self.ExportBtn.pack(side=LEFT, anchor='s', pady=10, ipady=4)

        # Кнопка "Удалить"
        self.DelBtn = ttk.Button(self.parent,
                                 text='Удалить',
                                 command=self.delete_selected,
                                 style='StyleRed.TButton')
        self.DelBtn.pack(side='right', anchor='s', padx=30, pady=10, ipadx=5, ipady=5)

        # Кнопка "Очистить всё"
        self.ClearBtn = ttk.Button(self.parent,
                                   text='Очистить всё',
                                   command=self.clear_all,
                                   style='StyleYellow.TButton')
        self.ClearBtn.pack(side='right', anchor='s', padx=10, pady=10, ipadx=5, ipady=5)

    def create_table(self):
        # Создаем фрейм для таблицы и скроллбара
        table_container = Frame(self.table_frame, background='white')
        table_container.pack(fill='both', expand=True, padx=10, pady=10)

        # Создаем вертикальный скроллбар
        scrollbar = Scrollbar(table_container)
        scrollbar.pack(side=RIGHT, fill=Y)

        # Создаем Treeview (таблицу)
        self.tree = ttk.Treeview(table_container,
                                 columns=('ID', 'User', 'Test', 'Total', 'Correct', 'Percent', 'Time'),
                                 show='headings',
                                 yscrollcommand=scrollbar.set,
                                 height=15)

        # Настраиваем скроллбар
        scrollbar.config(command=self.tree.yview)

        # Определяем заголовки колонок
        columns_config = [
            ('ID', 'ID', 50),
            ('User', 'Пользователь', 100),
            ('Test', 'Тест', 150),
            ('Total', 'Всего вопросов', 100),
            ('Correct', 'Правильно', 80),
            ('Percent', 'Процент', 80),
            ('Time', 'Время', 80),
        ]

        for col_id, heading, width in columns_config:
            self.tree.heading(col_id, text=heading)
            self.tree.column(col_id, width=width, anchor='center')

        # Упаковываем таблицу
        self.tree.pack(side=LEFT, fill='both', expand=True)

        # Бинд для выделения строк
        self.tree.bind('<<TreeviewSelect>>', self.on_select)

    def load_test_results(self):
        try:
            # Очищаем существующие данные
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Загружаем данные с API
            response = requests.get('http://localhost:8000/api/show_tests_results', timeout=10)

            if response.status_code == 200:
                data = response.json()
                results = data.get('results', [])

                if not results:
                    # Если нет данных, показываем сообщение
                    self.tree.insert('', 'end', values=('Нет данных', '', '', '', '', '', ''))
                    return

                # Заполняем таблицу данными
                for result in results:
                    # Форматируем время (секунды в мм:сс)
                    minutes, seconds = divmod(result.get('time_complete', 0), 60)
                    time_str = f"{minutes:02d}:{seconds:02d}"

                    self.tree.insert('', 'end',
                                     values=(
                                         result.get('id', ''),
                                         result.get('user_name', ''),
                                         result.get('test_name', ''),
                                         result.get('total_questions', ''),
                                         result.get('correct_answers', ''),
                                         f"{result.get('percent_correct_answers', 0):.1f}%",
                                         time_str,
                                     ))

            else:
                self.tree.insert('', 'end', values=('Ошибка загрузки', '', '', '', '', '', ''))

        except requests.exceptions.ConnectionError:
            self.tree.insert('', 'end', values=('Ошибка подключения', 'Запустите сервер', '', '', '', '', ''))
        except Exception as e:
            self.tree.insert('', 'end', values=(f'Ошибка: {str(e)}', '', '', '', '', '', ''))

    def on_select(self, event):
        """Обработчик выбора строк в таблице"""
        self.selected_items = self.tree.selection()

    def delete_selected(self):
        """Удаление выбранных записей"""
        if not self.selected_items:
            self.LableError.config(foreground='red')
            return

        result = messagebox.askyesno("Подтверждение", "Удалить выбранные записи?")

        if result:
            try:
                # Получаем ID выбранных записей
                ids_to_delete = []
                for item in self.selected_items:
                    item_values = self.tree.item(item)['values']
                    if item_values and item_values[0] != 'Нет данных':
                        # Преобразуем ID в число, если возможно
                        try:
                            ids_to_delete.append(int(item_values[0]))
                        except (ValueError, TypeError):
                            continue

                if not ids_to_delete:
                    messagebox.showwarning("Внимание", "Нечего удалять")
                    return

                # Отправляем запрос на удаление
                response = requests.delete(
                    'http://localhost:8000/api/delete_results',
                    json={'ids': ids_to_delete},
                    timeout=10
                )

                if response.status_code == 200:
                    result_data = response.json()
                    self.load_test_results()
                    messagebox.showinfo("Успех", f"Удалено записей: {result_data.get('deleted_count', 0)}")
                    self.LableError.config(foreground='white')
                else:
                    error_detail = response.json().get('detail', 'Unknown error')
                    messagebox.showerror("Ошибка", f"Ошибка сервера: {error_detail}")

            except requests.exceptions.ConnectionError:
                messagebox.showerror("Ошибка", "Не удалось подключиться к серверу")
            except requests.exceptions.Timeout:
                messagebox.showerror("Ошибка", "Превышено время ожидания сервера")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при удалении: {str(e)}")

    def clear_all(self):
        """Очистка всех записей"""
        result = messagebox.askyesno("Подтверждение", "Удалить ВСЕ записи? Это действие нельзя отменить!")

        if result:
            try:
                # Отправляем запрос на очистку всех данных
                response = requests.delete(
                    'http://localhost:8000/api/clear_all_results',
                    timeout=10
                )

                if response.status_code == 200:
                    result_data = response.json()
                    self.load_test_results()
                    messagebox.showinfo("Успех", f"Удалено всех записей: {result_data.get('deleted_count', 0)}")
                else:
                    error_detail = response.json().get('detail', 'Unknown error')
                    messagebox.showerror("Ошибка", f"Ошибка сервера: {error_detail}")

            except requests.exceptions.ConnectionError:
                messagebox.showerror("Ошибка", "Не удалось подключиться к серверу")
            except requests.exceptions.Timeout:
                messagebox.showerror("Ошибка", "Превышено время ожидания сервера")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при очистке: {str(e)}")

    def export_excel(self):
        """Экспорт результатов в Excel"""
        try:
            all_items = self.tree.get_children()
            if not all_items:
                messagebox.showwarning("Внимание", "Нет данных для экспорта")
                return

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Сохранить результат как"
            )

            if not file_path:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Результаты тестирований"

            # Стили для форматирования
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Заголовки столбцов
            headers = ['ID', 'Пользователь', 'Тест', 'Всего вопросов', 'Правильно', 'Процент', 'Время']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

                # Заполняем данными
                row_num = 2
                for item in all_items:
                    item_values = self.tree.item(item)['values']

                    # Пропускаем строки с сообщениями об ошибках или отсутствии данных
                    if not item_values:
                        continue

                    for col, value in enumerate(item_values, 1):
                        cell = ws.cell(row=row_num, column=col, value=value)
                        cell.border = border
                        if col in [1, 2]:  # Выравнивание по левому краю для текстовых полей
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="center", vertical="center")

                    row_num += 1

                    # Автоматическая ширина столбцов
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2)
                        ws.column_dimensions[column_letter].width = adjusted_width

                # Добавляем информацию о дате экспорта
                ws.cell(row=row_num + 1, column=1, value="Дата экспорта:").font = Font(bold=True)
                ws.cell(row=row_num + 1, column=2, value=datetime.now().strftime("%d.%m.%Y %H:%M"))

                ws.cell(row=row_num + 2, column=1, value="Всего записей:").font = Font(bold=True)
                ws.cell(row=row_num + 2, column=2, value=row_num - 2)

            # Сохраняем файл
            wb.save(file_path)

            messagebox.showinfo("Успех", f"Данные успешно экспортированы в файл:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Ошибка", f"Произошла ошибка при экспорте в Excel:\n{str(e)}")

    # Функция возвращения на экран главного меню
    def back_button(self):
        from pages.main_menu import MainMenu  # Импорт не в начале, т.к. иначе будет ошибка
        self.app.show_page(MainMenu)

    # Очистка виджетов при закрытии страницы
    def destroy(self):
        for widget in self.parent.winfo_children():
            widget.destroy()